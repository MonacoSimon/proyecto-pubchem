import requests
import pandas as pd
import tkinter as tk
from tkinter import messagebox
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

pug = "https://pubchem.ncbi.nlm.nih.gov/rest/pug"
ruta_excel = os.path.expanduser("/home/simonm/Escritorio/smiles/Compuestos.xlsx")

def limpiar_nombre(nombre):
    nombre = nombre.strip().lower()
    nombre = nombre.replace("-", " ").replace("_", " ")
    nombre = re.sub(r'\b[ivxlcdm]+\b', '', nombre)
    nombre = re.sub(r'\d+[o\-]', '', nombre)
    palabras_eliminar = ['derivative', 'analog', 'compound', 'extract', 'form', 'type']
    for palabra in palabras_eliminar:
        nombre = re.sub(r'\b' + palabra + r'\b', '', nombre)
    nombre = re.sub(r'\s+', ' ', nombre).strip()
    return nombre.title()

def buscar_alternativas_pubchem(nombre):
    try:
        url_busqueda = f"{pug}/compound/name/{nombre}/synonyms/TXT"
        response = requests.get(url_busqueda, timeout=10)
        if response.status_code == 200:
            sinonimos = response.text.strip().split('\n')
            return [sin for sin in sinonimos if sin and len(sin) < 100][:5]
    except:
        pass
    return []

def obtener_smiles_con_estereoquimia(cid):
    try:
        url = f"{pug}/compound/cid/{cid}/property/IsomericSMILES/TXT"
        response = requests.get(url, timeout=15)
        response.raise_for_status()
        smiles = response.text.strip()

        if '@' in smiles:
            return smiles
        
        url_todos = f"{pug}/compound/cid/{cid}/property/IsomericSMILES,CanonicalSMILES,SMILES/TXT"
        response_todos = requests.get(url_todos, timeout=15)
        if response_todos.status_code == 200:
            todos_smiles = response_todos.text.strip().split('\n')
            for smi in todos_smiles:
                if '@' in smi:
                    return smi
        
        url_json = f"{pug}/compound/cid/{cid}/JSON"
        response_json = requests.get(url_json, timeout=15)
        if response_json.status_code == 200:
            data = response_json.json()
            if 'PC_Compounds' in data and len(data['PC_Compounds']) > 0:
                compound = data['PC_Compounds'][0]
                if 'props' in compound:
                    for prop in compound['props']:
                        if (prop.get('urn', {}).get('label') == 'SMILES' and 
                            prop.get('urn', {}).get('name') == 'Isomeric' and
                            'value' in prop.get('value', {})):
                            smiles_val = prop['value']['value']
                            if '@' in smiles_val:
                                return smiles_val
        
        return smiles
        
    except Exception as e:
        print(f"Error obteniendo smile: {e}")
        return None

def obtener_smiles(compuestos):
    resultados = []
    errores = []
    
    for compuesto_original in compuestos:
        compuesto_original = compuesto_original.strip()
        if compuesto_original == "":
            continue
        
        compuesto_limpio = limpiar_nombre(compuesto_original)
        encontrado = False
        
        nombres_a_probar = [compuesto_original, compuesto_limpio]
        
        try:
            sinonimos = buscar_alternativas_pubchem(compuesto_limpio)
            nombres_a_probar.extend(sinonimos)
        except:
            pass
        
        for nombre_intento in nombres_a_probar:
            if not nombre_intento:
                continue
                
            try:
                print(f"Intentando con: {nombre_intento}")
                url_cid = f"{pug}/compound/name/{nombre_intento}/cids/TXT"
                cid_response = requests.get(url_cid, timeout=15)
                if cid_response.status_code == 404:
                    continue
                cid_response.raise_for_status()
                cid = cid_response.text.strip()
                if not cid:
                    continue
                smiles = obtener_smiles_con_estereoquimia(cid)
                
                if smiles:
                    resultados.append({
                        "Nombre": compuesto_original, 
                        "Smile": smiles
                    })
                    encontrado = True
                    break
                    
            except requests.exceptions.RequestException as e:
                print(f"Error con {nombre_intento}: {e}")
                continue
            except Exception as e:
                print(f"Error con {nombre_intento}: {e}")
                continue
        
        if not encontrado:
            errores.append(compuesto_original)
            print(f"Error: No se pudo encontrar {compuesto_original} en PubChem")
    
    return resultados, errores

def encontrar_fila_datos(ws):
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=2).value is None:
            return row
    return ws.max_row + 1

def procesar():
    entrada = text_entry.get("1.0", tk.END)
    compuestos = [c.strip() for c in entrada.split(",") if c.strip()]
    
    if not compuestos:
        messagebox.showwarning("Advertencia", "No se ingresaron compuestos.")
        return
    
    nuevos, errores = obtener_smiles(compuestos)
    
    try:
        wb = load_workbook(ruta_excel)
        ws = wb.active
        
        fila_inicio = encontrar_fila_datos(ws)
        
        font_8pt = Font(size=8)
        alignment_nowrap = Alignment(wrap_text=False, vertical='center')
        
        ws.column_dimensions['B'].width = 50  

        for i, compuesto in enumerate(nuevos):
            ws.cell(row=fila_inicio + i, column=1, value=compuesto["Nombre"])
            
            celda_smiles = ws.cell(row=fila_inicio + i, column=2, value=compuesto["Smile"])
            celda_smiles.font = font_8pt
            celda_smiles.alignment = alignment_nowrap
            
            ws.cell(row=fila_inicio + i, column=3, value=None)
        
        wb.save(ruta_excel)
        
        if errores:
            messagebox.showwarning("Completado parcialmente", 
                                  f"Se agregaron {len(nuevos)} compuestos correctamente.\n\n"
                                  f"Compuestos no encontrados ({len(errores)}):\n{', '.join(errores[:5])}"
                                  + (f"\n... y {len(errores)-5} más" if len(errores) > 5 else ""))
        else:
            messagebox.showinfo("Éxito", f"Se agregaron {len(nuevos)} compuestos a {ruta_excel} ✅")
        
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar el archivo: {e}")
        print(f"Error: {e}")

root = tk.Tk()
root.title("Agregar compuestos a Excel")
root.geometry("600x400")
root.resizable(False, False)

label = tk.Label(root, text="Escriba los compuestos separados por comas:")
label.pack(pady=10)

text_entry = tk.Text(root, height=15, width=70)
text_entry.pack(pady=10)

boton = tk.Button(root, text="Aceptar", command=procesar, width=20, height=2)
boton.pack(pady=10)

root.mainloop()